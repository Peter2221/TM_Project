from openpyxl import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell

class ExcelHandler():
    produkty = []
    kalorie = []

    def open_workbook(self, filename):
        wb = load_workbook(filename, data_only=True)  # wybieramy plik
        return wb

    #nie jestem pewna czy ta funkcja jest potrzebna ale napisałam :3
    def assign_sheets(self, wb):
        self.produkty = wb["produkty"]  # wybieramy arkusz
        self.kalorie = wb["kalorie"]

    def find_product(self, product):

        #while True:
        #searched_product = input("podaj produkt: ") #to zastapimy mowa + morfeusz
        searched_product = product
        found_cell = None

        for row in range(1, self.produkty.max_row + 1):
            for col in range(1, self.produkty.max_column + 1):
                my_cell = self.produkty.cell(row, col)
                if my_cell.value == searched_product:
                    found_cell = xl_rowcol_to_cell(row - 1, col - 1)
                    return found_cell

        if found_cell == None:
            raise NoProductException("Nie ma takiego produktu w bazie!")


    def get_calories(self, found_cell):
        kcal = self.kalorie[found_cell]
        return kcal

    def what_you_ate(self, product, product_weight):
        found_cell = self.find_product(product)
        product = self.produkty[found_cell]
        kcal = self.get_calories(found_cell)
        kcal = product_weight/100 * kcal.value
        print("zjadłeś produkt: " + str(product.value) + " co daje: " + str(kcal) + " kcal")


class NoProductException(Exception):
    pass














